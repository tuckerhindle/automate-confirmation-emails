from datetime import datetime
from openpyxl import load_workbook

import win32com.client as win32


def email_content(name, term, course, section):
    '''Takes a student name and course information, returns a subject line and body of an email.'''

    subject = f'Approved to Register for {term} {course}'

    greeting = f'Hello {name},<br><br>'
    message = f'Thank you for completing the {term} Research Courses form.  The instructor approved your request for <b>{course} at {section}</b>.<br><br>Once FAU registration opens, please add this class to your schedule by following these instructions: <a href="https://www.fau.edu/registrar/registration/#Register">FAU Guide to Register for Classes</a><br><br>'
    signature = f'We are excited to see you this {term}!<br><br>--<br><b>FAU Lab Schools Research Team</b><br>Florida Atlantic University<br>777 Glades Road, Building 26F, Room 109<br>Boca Raton, Florida 33431'

    body = '<h3 style="font-weight:normal;">' + greeting + message + signature + '</h3>'

    return subject, body


def send_new_email(subject, recipients, body):
    '''Sends a new email using MS Outlook to the specified recipients containing the subject line and body provided.'''
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)

    mail.Subject = subject
    mail.To = recipients
    mail.HTMLBody = body

    mail.Send()

# User inputs
term = 'Fall 2023'
workbook = 'Fall 2023 Course Request Form.xlsx'
names = 7
emails = 8
courses = 11
status_col = 15
updated = 16

# Send email to approved course registrations only, update spreadsheet
wb = load_workbook(workbook)
ws = wb.active

row_count = ws.max_row

for row in range(2, row_count + 1):

    status = ws.cell(row=row, column=status_col).value
    course1 = ws.cell(row=row, column=12).value
    course2 = ws.cell(row=row, column=13).value

    if status == 'Approved':
        name = ws.cell(row=row, column=names).value
        recipient = ws.cell(row=row, column=emails).value
        course = ws.cell(row=row, column=courses).value
        section = course2 if course1 is None else course1
        
        subject, body = email_content(name, term, course, section)

        send_new_email(subject, recipient, body)

        new_status = 'done'
        ws.cell(row=row, column=status_col).value = new_status
        ws.cell(row=row, column=updated).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

wb.save(workbook)

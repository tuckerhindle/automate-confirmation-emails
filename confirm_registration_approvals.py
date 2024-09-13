from openpyxl import load_workbook
from datetime import datetime

from outlook_utils import get_cell_value, write_email_content, send_new_email

# User Inputs
workbook = r"C:\Users\thindle2016\Florida Atlantic University\FAU Lab Schools Research Department - Documents\FAUHS Research Program\Research Courses\Forms\Course signup forms\Spring_2025_Registration_Materials\Spring 2025 Research Courses Form.xlsx"
term = "Spring 2025"
outlook_account = "fauhsresearch@fau.edu"

# Specify column index
names = 8
emails = 9
courses = 12
EDF2911_sections = 13
EDF3913_sections = 14
status_col = 16
last_updated = 17

wb = load_workbook(workbook)
ws = wb.active

row_count = ws.max_row

for row in range(2, row_count + 1):

    status = get_cell_value(ws, row, status_col)

    if status == "Approved":
        name = get_cell_value(ws, row, names)
        recipient = get_cell_value(ws, row, emails)
        course = get_cell_value(ws, row, courses)

        EDF2911 = get_cell_value(ws, row, EDF2911_sections) # course 1
        EDF3913 = get_cell_value(ws, row, EDF3913_sections) # course 2
        section = EDF3913 if EDF2911 is None else EDF2911
        
        subject, body = write_email_content(name, term, course, section)

        send_new_email(outlook_account, recipient, subject, body)

        new_status = "Email Sent"
        ws.cell(row=row, column=status_col).value = new_status
        ws.cell(row=row, column=last_updated).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

wb.save(workbook)
